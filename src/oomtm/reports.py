"""oomtm.reports — aktliste / sagsliste renderers (Excel + PDF).

Pure renderers: they take already-prepared row dicts and return file *bytes*.
No SharePoint / GO / network here — the caller (an OO robot) uploads the bytes.

Ported from the proven AktBob aktliste generators
(``Python_Aktbob2_FromFilarkivToSharePoint/GenerateAndUploadAktliste.py``), minus
pandas, plus two fixes: the PDF column header now repeats on every page
(``repeatRows=1`` + a later-page template) and cell text is XML-escaped so a
title containing ``&`` / ``<`` no longer breaks the PDF.

``openpyxl`` and ``reportlab`` are the ``oomtm[reports]`` extra.
"""
from __future__ import annotations

import io
import textwrap
import xml.sax.saxutils as _su

# Internal row-key -> column header, in display order. The robot builds rows as
# dicts keyed by the left value; both the Excel and PDF use this single list so
# the two formats can never drift apart.
AKTLISTE_COLUMNS: list[tuple[str, str]] = [
    ("akt_id", "Akt ID"),
    ("filnavn", "Filnavn"),
    ("kategori", "Kategori"),
    ("dato", "Dato"),
    ("dok_id", "Dok ID"),
    ("bilag_til", "Bilag til Dok ID"),
    ("bilag", "Bilag"),
    ("omfattet", "Omfattet af aktindsigt?"),
    ("gives", "Gives der aktindsigt?"),
    ("begrundelse", "Begrundelse"),
]


def _str(value) -> str:
    return "" if value is None else str(value)


def _compact_attachments(value, max_visible: int = 5) -> str:
    """'a, b, c, …' but collapse a long bilag list to '…i alt N bilag'."""
    text = _str(value).replace("\n", " ").replace("\r", " ").strip()
    if not text:
        return ""
    parts = [p.strip() for p in text.split(",") if p.strip()]
    if len(parts) <= max_visible:
        return ", ".join(parts)
    return f"{', '.join(parts[:max_visible])} (...) i alt {len(parts)} bilag"


def _wrap_markup(text, max_chars: int) -> str:
    """XML-escape ``text`` and soft-wrap it to ~max_chars per line with <br/>.

    reportlab Paragraph parses its input as markup, so unescaped ``&``/``<``/``>``
    would raise — escape first, then insert <br/> (which Paragraph treats as a
    line break). Paragraph also wraps long unbroken tokens to the column width on
    its own, so this is only a hint for nicer breaks."""
    if text is None or text == "":
        return ""
    escaped = _su.escape(str(text))
    lines: list[str] = []
    line = ""
    for word in escaped.split():
        if len(line) + len(word) + 1 <= max_chars:
            line = f"{line} {word}".strip()
        else:
            if line:
                lines.append(line)
            line = word
    if line:
        lines.append(line)
    return "<br/>".join(lines)


def aktliste_xlsx(rows: list[dict]) -> bytes:
    """Render the aktliste as a styled .xlsx (sheet 'Aktliste') and return bytes.

    ``rows`` is a list of dicts keyed by :data:`AKTLISTE_COLUMNS` keys."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    keys = [k for k, _ in AKTLISTE_COLUMNS]
    headers = [h for _, h in AKTLISTE_COLUMNS]

    wb = Workbook()
    ws = wb.active
    ws.title = "Aktliste"

    ws.row_dimensions[1].height = 20
    for ci, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        letter = get_column_letter(ci)
        if header == "Filnavn":
            ws.column_dimensions[letter].width = 79
        elif header == "Begrundelse":
            ws.column_dimensions[letter].width = 40
        else:
            ws.column_dimensions[letter].width = max(15, len(header) + 5)

    for ri, row in enumerate(rows, start=2):
        max_lines = 1
        for ci, key in enumerate(keys, start=1):
            value = row.get(key)
            cell = ws.cell(row=ri, column=ci, value=_str(value))
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if key == "filnavn" and value:
                max_lines = max(max_lines, len(textwrap.wrap(str(value), width=70)) or 1)
        ws.row_dimensions[ri].height = 15 * max_lines

    # An Excel table needs at least one data row; skip the styled table for an
    # empty aktliste (still a valid sheet with just the header).
    if rows:
        last_col = get_column_letter(len(headers))
        table = Table(displayName="AktTable", ref=f"A1:{last_col}{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def aktliste_pdf(rows: list[dict], *, sagsnummer: str, dato_string: str,
                 logo_path: str | None = None) -> bytes:
    """Render the aktliste as a landscape-A4 PDF and return bytes.

    Header band (page 1): optional AAK logo, ``Aktliste - {sagsnummer}`` title,
    and ``Dato for aktindsigt: {dato_string}``. The blue column-header row repeats
    on every page."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import (Frame, NextPageTemplate, PageTemplate,
                                    Paragraph, SimpleDocTemplate)
    from reportlab.platypus import Table as RTable
    from reportlab.platypus import TableStyle

    page_width, page_height = landscape(A4)
    margin = 40
    base = getSampleStyleSheet()["Normal"]
    header_style = ParagraphStyle("akt_header", parent=base, fontName="Helvetica-Bold",
                                  fontSize=10, textColor=colors.white, alignment=1, leading=12)
    cell_style = ParagraphStyle("akt_cell", parent=base, fontName="Helvetica",
                                fontSize=8, textColor=colors.black, alignment=1, leading=10)

    keys = [k for k, _ in AKTLISTE_COLUMNS]
    headers = [h for _, h in AKTLISTE_COLUMNS]
    column_widths = [50, 150, 80, 70, 75, 55, 50, 65, 70, 100]
    char_limits = [10, 30, 15, 12, 15, 10, 9, 12, 12, 20]

    table_data = [[Paragraph(h, header_style) for h in headers]]
    for row in rows:
        cells = []
        for i, key in enumerate(keys):
            raw = row.get(key)
            if key == "bilag":
                raw = _compact_attachments(raw)
            cells.append(Paragraph(_wrap_markup(raw, char_limits[i]), cell_style))
        table_data.append(cells)

    table = RTable(table_data, colWidths=column_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3661D8")),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))

    def draw_header(canvas, _doc):
        canvas.saveState()
        img_w, img_h = 100, 45
        img_y = page_height - margin - img_h
        if logo_path:
            try:
                canvas.drawImage(logo_path, margin, img_y, width=img_w, height=img_h,
                                 preserveAspectRatio=True, mask="auto")
            except Exception:  # pylint: disable=broad-except
                pass  # logo is decorative — never fail the aktliste over it
        title = f"Aktliste - {sagsnummer}"
        canvas.setFont("Helvetica-Bold", 14)
        title_y = img_y - 20
        canvas.drawString(margin, title_y, title)
        title_w = canvas.stringWidth(title, "Helvetica-Bold", 14)
        canvas.setStrokeColor(colors.black)
        canvas.setLineWidth(1)
        canvas.line(margin, title_y - 5, margin + title_w, title_y - 5)
        date_string = f"Dato for aktindsigt: {dato_string}"
        canvas.setFont("Helvetica", 10)
        date_w = canvas.stringWidth(date_string, "Helvetica", 10)
        canvas.drawString(page_width - margin - date_w, img_y, date_string)
        canvas.restoreState()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=margin,
                            rightMargin=margin, topMargin=margin, bottomMargin=margin)
    # Page 1 reserves a band at the top for the logo/title/date; later pages use
    # the full height (the column header repeats via repeatRows).
    first_h = (page_height - margin - 100) - margin
    frame_first = Frame(margin, margin, page_width - 2 * margin, first_h, id="first")
    frame_later = Frame(margin, margin, page_width - 2 * margin, page_height - 2 * margin, id="later")
    doc.addPageTemplates([
        PageTemplate(id="First", frames=frame_first, onPage=draw_header),
        PageTemplate(id="Later", frames=frame_later),
    ])
    doc.build([NextPageTemplate("Later"), table])
    return buf.getvalue()
